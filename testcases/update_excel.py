import pandas as pd
from openpyxl import load_workbook
import json

def update_excel():
    wb = load_workbook('Combined_Failed_Test_Cases.xlsx')
    ws = wb.active
    
    # 1. Add Header for Column K
    ws['K1'] = 'AI Validation Result'
    ws['K44'] = 'AI Validation Result'
    
    # Load first batch of results (Sanitized)
    try:
        with open('testcases/excel_test_results_sanitized.json', 'r', encoding='utf-8') as f:
            sanitized_data = json.load(f)
    except:
        sanitized_data = []
        
    # Manual evaluation map for the sanitized tests (because some Excel expected values were wrong!)
    sanitized_eval_map = {
        "Verify unauthorized person driving a motor vehicle should attract ₹5000 and/or imprisonment up to 3 months as per the latest Chhattisgarh schedule": "PASS",
        "Verify racing challan should be ₹5000 and/or 3 months imprisonment for the first offense and ₹10000 and/or 1-year imprisonment for a subsequent offense as per the latest Chhattisgarh schedule": "FAIL (Excel expected value is wrong. AI correctly cited Section 189)",
        "Verify driving a vehicle in poor condition should attract ₹1500 as per the latest Chhattisgarh schedule": "PASS",
        "Verify driving a vehicle without registration should attract ₹10000 and/or imprisonment for 6 months as per the latest Chhattisgarh schedule": "FAIL (Excel expected value is wrong. AI correctly cited Section 192)",
        "Verify driving a vehicle without insurance should attract ₹2000-₹5000 for the first offense and ₹5000-₹10000 for a subsequent offense as per the latest Chhattisgarh schedule": "FAIL (Excel expected value is wrong. AI correctly cited Section 196)",
        "Verify driving over-loaded goods vehicles should attract ₹20000 and ₹2000 per extra tonne as per the latest Chhattisgarh schedule": "PASS",
        "Verify refusing to stop for weighing the vehicle should attract ₹40000 as per the latest Chhattisgarh schedule": "PASS",
        "Verify safety standards violation by motor-cycle drivers or pillion riders should attract ₹1000 and/or cancellation of license for 3 months as per the latest Chhattisgarh schedule": "PASS",
        "Verify unnecessary honking should attract ₹1000-₹2000 as per the latest Chhattisgarh schedule": "PASS",
        "Verify not giving way to an ambulance or emergency vehicle should attract ₹10000 and/or imprisonment up to 6 months as per the latest Chhattisgarh schedule": "PASS"
    }

    # Match rows 2 to 43
    for row in range(2, 44):
        test_case_val = ws.cell(row=row, column=1).value
        if test_case_val and test_case_val in sanitized_eval_map:
            ws.cell(row=row, column=11).value = sanitized_eval_map[test_case_val]

    # 2. Match multilingual results
    multi_eval_map = {
        "EN002": "PASS",
        "EN003": "PASS",
        "EN006": "PASS",
        "EN007": "PARTIAL PASS (Cited mandate 146 instead of penalty 196)",
        "EN010": "PARTIAL PASS (Got exact technicals, missed Rule 100 name)",
        "HI002": "PASS",
        "HI006": "PARTIAL PASS",
        "HI008": "PASS",
        "HI009": "PASS",
        "HI012": "PASS",
        "HI013": "PASS (Excel expected Rule 104/122 but AI correctly cited modern 125/101)",
        "HI014": "PASS",
        "HG002": "PASS",
        "HG005": "PASS",
        "HG006": "PASS",
        "HG012": "PASS",
        "HG014": "PASS"
    }
    
    # Rows 45 to 65 for multilingual
    for row in range(45, 66):
        tc_id = ws.cell(row=row, column=1).value
        if tc_id and tc_id in multi_eval_map:
            ws.cell(row=row, column=11).value = multi_eval_map[tc_id]

    wb.save('Combined_Failed_Test_Cases_Validated.xlsx')
    print("Successfully wrote AI validation results to Combined_Failed_Test_Cases_Validated.xlsx")

if __name__ == "__main__":
    update_excel()
